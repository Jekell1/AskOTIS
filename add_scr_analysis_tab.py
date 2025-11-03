"""
Add a new tab to program_scn_relationships.xlsx with analysis of files containing "SCR" in their names.
For each file, extract description from file headers and comments.
"""

import os
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re

def find_scr_files(base_path):
    """Find all files with 'SCR' in their names in the cobol_src directory."""
    scr_files = []
    
    # Search for all files with SCR in the name
    for root, dirs, files in os.walk(base_path):
        for file in files:
            if 'SCR' in file.upper():
                full_path = os.path.join(root, file)
                # Get relative path from cobol_src
                rel_path = os.path.relpath(full_path, base_path)
                scr_files.append({
                    'filename': file,
                    'path': rel_path,
                    'full_path': full_path
                })
    
    return sorted(scr_files, key=lambda x: x['filename'])

def read_file_content(file_path, max_chars=50000):
    """Read file content, handling potential encoding issues."""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read(max_chars)
        return content
    except Exception as e:
        return f"Error reading file: {str(e)}"

def get_openai_description(file_path, file_content):
    """Get description from OpenAI about what the file does."""
    try:
        # Truncate content if too long (keep first 30000 chars)
        if len(file_content) > 30000:
            file_content = file_content[:30000] + "\n... [truncated]"
        
        prompt = f"""Analyze this COBOL source file and provide a concise description (2-3 sentences) of what it does.
Focus on the main purpose and functionality.

File: {file_path}

Content:
{file_content}
"""
        
        response = client.chat.completions.create(
            model="gpt-4",  # Use your Azure deployment name
            messages=[
                {"role": "system", "content": "You are a COBOL expert analyzing legacy code. Provide clear, concise descriptions."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=300,
            temperature=0.3
        )
        
        description = response.choices[0].message.content.strip()
        return description
        
    except Exception as e:
        return f"Error getting OpenAI description: {str(e)}"

def add_scr_tab_to_excel(excel_path, scr_files_data):
    """Add a new tab with SCR file analysis to the existing Excel file."""
    
    # Load existing workbook
    wb = openpyxl.load_workbook(excel_path)
    
    # Create new sheet (or replace if exists)
    if 'SCR Files Analysis' in wb.sheetnames:
        del wb['SCR Files Analysis']
    
    ws = wb.create_sheet('SCR Files Analysis', 2)  # Insert as 3rd sheet
    
    # Define headers
    headers = ['File Name', 'File Path', 'Description (OpenAI Analysis)']
    
    # Write headers with formatting
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 30  # File Name
    ws.column_dimensions['B'].width = 40  # File Path
    ws.column_dimensions['C'].width = 80  # Description
    
    # Write data
    for row_idx, file_data in enumerate(scr_files_data, start=2):
        ws.cell(row=row_idx, column=1, value=file_data['filename'])
        ws.cell(row=row_idx, column=2, value=file_data['path'])
        ws.cell(row=row_idx, column=3, value=file_data['description'])
        
        # Wrap text for description
        ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Enable auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Save workbook
    wb.save(excel_path)
    print(f"\nExcel file updated: {excel_path}")

def main():
    """Main function to process SCR files and update Excel."""
    
    print("=" * 70)
    print("SCR FILES ANALYSIS - Adding Tab to Excel Report")
    print("=" * 70)
    
    base_path = os.path.join(os.getcwd(), 'cobol_src')
    excel_path = os.path.join(os.getcwd(), 'program_scn_relationships.xlsx')
    
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        return
    
    # Find all SCR files
    print("\nSearching for files with 'SCR' in their names...")
    scr_files = find_scr_files(base_path)
    print(f"Found {len(scr_files)} files with 'SCR' in their names")
    
    if not scr_files:
        print("No SCR files found!")
        return
    
    # Process each file
    scr_files_data = []
    
    for idx, file_info in enumerate(scr_files, start=1):
        print(f"\nProcessing {idx}/{len(scr_files)}: {file_info['filename']}")
        
        # Read file content
        content = read_file_content(file_info['full_path'])
        
        # Get OpenAI description
        print(f"  Getting OpenAI analysis...")
        description = get_openai_description(file_info['path'], content)
        
        scr_files_data.append({
            'filename': file_info['filename'],
            'path': file_info['path'],
            'description': description
        })
        
        # Show preview of description
        preview = description[:100] + "..." if len(description) > 100 else description
        print(f"  Description: {preview}")
        
        # Small delay to avoid rate limiting
        time.sleep(0.5)
    
    # Add to Excel
    print("\n" + "=" * 70)
    print("Adding SCR analysis tab to Excel...")
    add_scr_tab_to_excel(excel_path, scr_files_data)
    
    # Summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"Total SCR files processed: {len(scr_files_data)}")
    print(f"Excel file updated: {excel_path}")
    print(f"New tab: 'SCR Files Analysis'")
    print("=" * 70)

if __name__ == "__main__":
    main()
