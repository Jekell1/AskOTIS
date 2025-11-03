"""
Create Excel spreadsheet of all program-to-SCN file relationships.
Includes program descriptions, SCN descriptions, and extracted screen text.
"""

import os
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def extract_description_from_cobol(file_path):
    """Extract description from COBOL file comments."""
    try:
        with open(file_path, 'r', encoding='latin-1', errors='ignore') as f:
            content = f.read(2000)  # Read first 2000 chars
            
            # Look for DESC: or DESCRIPTION: patterns
            desc_patterns = [
                r'(?:DESC|DESCRIPTION)\s*:\s*(.+?)(?:\n|$)',
                r'\*\s+(.+?)\s+\*',
                r'PROGRAM-ID\.\s+\w+\.\s*\*\s*(.+?)(?:\n|$)'
            ]
            
            for pattern in desc_patterns:
                match = re.search(pattern, content, re.IGNORECASE | re.MULTILINE)
                if match:
                    desc = match.group(1).strip()
                    # Clean up
                    desc = re.sub(r'\s+', ' ', desc)
                    desc = desc.replace('*', '').strip()
                    if len(desc) > 10:  # Meaningful description
                        return desc[:200]  # Limit length
    except Exception as e:
        pass
    
    return ""

def extract_screen_text_from_scn(file_path):
    """Extract screen text from SCN copybook file."""
    screen_texts = []
    
    try:
        with open(file_path, 'r', encoding='latin-1', errors='ignore') as f:
            for line in f:
                # Look for LABEL LINE with text
                label_match = re.search(r'LABEL\s+LINE\s+\d+\s+.*?"([^"]+)"', line, re.IGNORECASE)
                if label_match:
                    text = label_match.group(1).strip()
                    if text and len(text) > 0:
                        screen_texts.append(text)
                
                # Look for VALUE clauses with text
                value_match = re.search(r'VALUE\s+"([^"]+)"', line, re.IGNORECASE)
                if value_match:
                    text = value_match.group(1).strip()
                    if text and len(text) > 0 and text not in screen_texts:
                        screen_texts.append(text)
                
                # Look for PIC X() VALUE patterns
                pic_value_match = re.search(r'PIC\s+X\(\d+\)\s+VALUE\s+"([^"]+)"', line, re.IGNORECASE)
                if pic_value_match:
                    text = pic_value_match.group(1).strip()
                    if text and len(text) > 0 and text not in screen_texts:
                        screen_texts.append(text)
    except Exception as e:
        pass
    
    # Return unique screen texts, joined with semicolons
    return " | ".join(screen_texts[:20]) if screen_texts else ""  # Limit to first 20 items

def find_all_scn_relationships():
    """Find all program-to-SCN relationships."""
    cobol_src = r"c:\Users\jeff.childers\Documents\OTISCodeResearcher\cobol_src"
    
    # Pattern to match COPY statements with SCN in the filename
    copy_pattern = re.compile(r'^\s*COPY\s+"([^"]*SCN[^"]*\.CPY)"', re.IGNORECASE)
    
    relationships = []
    
    print("Scanning COBOL programs for SCN usage...")
    file_count = 0
    
    # Walk through all .CBL files
    for root, dirs, files in os.walk(cobol_src):
        for file in files:
            if file.upper().endswith('.CBL'):
                file_count += 1
                if file_count % 100 == 0:
                    print(f"  Processed {file_count} programs...")
                
                program_path = os.path.join(root, file)
                rel_program_path = os.path.relpath(program_path, cobol_src)
                
                # Get program description
                program_desc = extract_description_from_cobol(program_path)
                
                try:
                    with open(program_path, 'r', encoding='latin-1', errors='ignore') as f:
                        for line_num, line in enumerate(f, 1):
                            match = copy_pattern.match(line)
                            if match:
                                scn_file = match.group(1)
                                scn_file_upper = scn_file.upper()
                                
                                # Try to find the actual SCN file
                                scn_full_path = None
                                scn_desc = ""
                                screen_text = ""
                                
                                # Look in the library directories
                                scn_parts = scn_file_upper.split('/')
                                if len(scn_parts) == 2:
                                    lib_dir = scn_parts[0]
                                    scn_name = scn_parts[1]
                                    potential_path = os.path.join(cobol_src, lib_dir, scn_name)
                                    if os.path.exists(potential_path):
                                        scn_full_path = potential_path
                                
                                # If not found, search for it
                                if not scn_full_path:
                                    for search_root, search_dirs, search_files in os.walk(cobol_src):
                                        for search_file in search_files:
                                            if search_file.upper() == scn_parts[-1]:
                                                scn_full_path = os.path.join(search_root, search_file)
                                                break
                                        if scn_full_path:
                                            break
                                
                                # Extract description and screen text if found
                                if scn_full_path:
                                    scn_desc = extract_description_from_cobol(scn_full_path)
                                    screen_text = extract_screen_text_from_scn(scn_full_path)
                                
                                # Add relationship
                                relationships.append({
                                    'program': file,
                                    'program_path': rel_program_path,
                                    'program_desc': program_desc,
                                    'scn_file': scn_file_upper,
                                    'scn_desc': scn_desc,
                                    'screen_text': screen_text,
                                    'line_num': line_num
                                })
                
                except Exception as e:
                    print(f"Error processing {rel_program_path}: {e}")
    
    print(f"\nTotal programs scanned: {file_count}")
    print(f"Total program-SCN relationships found: {len(relationships)}")
    
    return relationships

def create_excel_report(relationships):
    """Create Excel workbook with all relationships."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Program-SCN Relationships"
    
    # Define headers
    headers = [
        'Program Name',
        'Program Path',
        'Program Description',
        'SCN Copybook',
        'SCN Description',
        'Screen Text',
        'Line Number'
    ]
    
    # Style definitions
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 20  # Program Name
    ws.column_dimensions['B'].width = 25  # Program Path
    ws.column_dimensions['C'].width = 40  # Program Description
    ws.column_dimensions['D'].width = 30  # SCN Copybook
    ws.column_dimensions['E'].width = 40  # SCN Description
    ws.column_dimensions['F'].width = 60  # Screen Text
    ws.column_dimensions['G'].width = 10  # Line Number
    
    # Sort relationships by program name, then SCN file
    relationships.sort(key=lambda x: (x['program'], x['scn_file']))
    
    # Write data rows
    for row_num, rel in enumerate(relationships, 2):
        ws.cell(row=row_num, column=1, value=rel['program']).alignment = wrap_alignment
        ws.cell(row=row_num, column=2, value=rel['program_path']).alignment = wrap_alignment
        ws.cell(row=row_num, column=3, value=rel['program_desc']).alignment = wrap_alignment
        ws.cell(row=row_num, column=4, value=rel['scn_file']).alignment = wrap_alignment
        ws.cell(row=row_num, column=5, value=rel['scn_desc']).alignment = wrap_alignment
        ws.cell(row=row_num, column=6, value=rel['screen_text']).alignment = wrap_alignment
        ws.cell(row=row_num, column=7, value=rel['line_num']).alignment = Alignment(horizontal='center')
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Add auto-filter
    ws.auto_filter.ref = f'A1:G{len(relationships) + 1}'
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15
    
    summary_data = [
        ['Metric', 'Count'],
        ['Total Relationships', len(relationships)],
        ['Unique Programs', len(set(r['program'] for r in relationships))],
        ['Unique SCN Files', len(set(r['scn_file'] for r in relationships))],
    ]
    
    for row_num, (label, value) in enumerate(summary_data, 1):
        cell_a = ws_summary.cell(row=row_num, column=1, value=label)
        cell_b = ws_summary.cell(row=row_num, column=2, value=value)
        
        if row_num == 1:
            cell_a.fill = header_fill
            cell_a.font = header_font
            cell_b.fill = header_fill
            cell_b.font = header_font
        else:
            cell_a.font = Font(bold=True)
    
    # Save workbook
    output_file = 'program_scn_relationships.xlsx'
    wb.save(output_file)
    print(f"\nExcel report saved to: {output_file}")
    print(f"Total rows: {len(relationships) + 1} (including header)")
    
    return output_file

def main():
    print("=" * 80)
    print("CREATING PROGRAM-SCN RELATIONSHIPS EXCEL REPORT")
    print("=" * 80)
    print()
    
    # Find all relationships
    relationships = find_all_scn_relationships()
    
    if not relationships:
        print("\nNo relationships found!")
        return
    
    # Create Excel report
    print("\nCreating Excel workbook...")
    output_file = create_excel_report(relationships)
    
    # Print sample
    print("\nSample of first 5 relationships:")
    print("-" * 80)
    for i, rel in enumerate(relationships[:5], 1):
        print(f"\n{i}. {rel['program']} -> {rel['scn_file']}")
        if rel['program_desc']:
            print(f"   Program: {rel['program_desc'][:60]}...")
        if rel['scn_desc']:
            print(f"   SCN: {rel['scn_desc'][:60]}...")
        if rel['screen_text']:
            print(f"   Text: {rel['screen_text'][:60]}...")
    
    print("\n" + "=" * 80)
    print("COMPLETE!")
    print("=" * 80)

if __name__ == "__main__":
    main()
