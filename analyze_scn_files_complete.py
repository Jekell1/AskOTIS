"""
Process all _SCN files to extract:
1. Basic description
2. Menu/screen information (options, fields, function keys)
3. Programs that call the _SCN file
Store results in both Excel and JSON formats.
"""

import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openai import AzureOpenAI
import time

def load_config():
    """Load configuration from local.settings.json file."""
    config_path = os.path.join(os.getcwd(), 'local.settings.json')
    
    if not os.path.exists(config_path):
        print(f"Error: Configuration file not found: {config_path}")
        return None
    
    try:
        with open(config_path, 'r') as f:
            settings = json.load(f)
            values = settings.get('Values', {})
            
            config = {
                'AZURE_OPENAI_API_KEY': values.get('AZURE_OPENAI_KEY') or values.get('AZURE_OPENAI_API_KEY'),
                'AZURE_OPENAI_ENDPOINT': values.get('AZURE_OPENAI_ENDPOINT'),
                'AZURE_OPENAI_DEPLOYMENT': values.get('AZURE_OPENAI_DEPLOYMENT', 'gpt-4')
            }
            return config
    except Exception as e:
        print(f"Error loading config file: {str(e)}")
        return None

def find_scn_files(base_path):
    """Find all files with '_SCN' in their names."""
    scn_files = []
    
    for root, dirs, files in os.walk(base_path):
        for file in files:
            if '_SCN' in file.upper():
                full_path = os.path.join(root, file)
                rel_path = os.path.relpath(full_path, base_path)
                scn_files.append({
                    'filename': file,
                    'path': rel_path,
                    'full_path': full_path
                })
    
    return sorted(scn_files, key=lambda x: x['filename'])

def read_file_content(file_path, max_chars=50000):
    """Read file content."""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read(max_chars)
        return content
    except Exception as e:
        return f"Error reading file: {str(e)}"

def find_calling_programs(scn_filename, base_path):
    """Find all programs that COPY this SCN file."""
    calling_programs = []
    scn_name = scn_filename.replace('.CPY', '').upper()
    
    # Search all .CBL files for COPY statements
    for root, dirs, files in os.walk(base_path):
        for file in files:
            if file.upper().endswith('.CBL'):
                file_path = os.path.join(root, file)
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                        # Look for COPY "path/SCN_NAME" or COPY SCN_NAME
                        if f'COPY "{scn_name}"' in content.upper() or \
                           f'COPY {scn_name}' in content.upper() or \
                           f'/{scn_filename.upper()}' in content.upper():
                            calling_programs.append(file)
                except:
                    pass
    
    return sorted(list(set(calling_programs)))

def extract_scn_analysis(client, deployment_name, file_path, file_content):
    """Get structured analysis from OpenAI."""
    try:
        if len(file_content) > 40000:
            file_content = file_content[:40000] + "\n... [truncated]"
        
        prompt = f"""Analyze this COBOL screen definition file (_SCN.CPY) and provide a structured analysis in the following format:

**CRITICAL INSTRUCTIONS:**
1. Extract EVERY menu option - do NOT stop at 10. If there are 15+ options, list ALL of them (1, 2, 3... 15, 16, etc.)
2. Extract ALL function keys (F1, F2, F3... F12) with their exact descriptions
3. Count menu options carefully - include EVERY numbered item you find
4. Look for VALUE clauses and literal strings to find complete menu text
5. Do NOT truncate or summarize - provide COMPLETE lists

**DESCRIPTION:**
Provide a clear 2-3 sentence description of what this screen does.

**MENU/SCREEN INFORMATION:**
List the following with COMPLETE information:
- Menu Options: If this is a menu screen, list EVERY SINGLE numbered option with descriptions (e.g., "1. DAILY PROCESSING", "2. REPORTS"... "11. ALTERNATE BRANCH LOGIN", "12. SYSTEM UTILITIES"). DO NOT STOP AT 10 - continue until all options are listed.
- Screen Fields: List key input/output fields if it's a data entry screen
- Function Keys: List ALL function keys you find (e.g., "F2: SIGN OFF", "F3: EXIT", "F7: MASTER MENU"). Include the exact text shown for each key.
- Screen Layout: Describe the general layout (columns, lines, sections)

**PURPOSE:**
Describe the role of this screen in the system and what business function it supports.

File: {file_path}

Content:
{file_content}

IMPORTANT: Extract EVERY menu option and EVERY function key. Do not truncate lists or skip items.
Provide the analysis in the exact format shown above with clear headings.
"""
        
        response = client.chat.completions.create(
            model=deployment_name,
            messages=[
                {"role": "system", "content": "You are a COBOL expert analyzing screen definition files. Provide structured, clear analysis following the exact format requested. Extract EVERY menu option and EVERY function key - do not truncate lists."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=2500,
            temperature=0.3
        )
        
        return response.choices[0].message.content.strip()
        
    except Exception as e:
        return f"Error getting analysis: {str(e)}"

def parse_analysis(analysis_text):
    """Parse the structured analysis into components."""
    result = {
        'description': '',
        'menu_screen_info': '',
        'purpose': ''
    }
    
    sections = {
        'DESCRIPTION:': 'description',
        'MENU/SCREEN INFORMATION:': 'menu_screen_info',
        'PURPOSE:': 'purpose'
    }
    
    current_section = None
    lines = analysis_text.split('\n')
    
    for line in lines:
        line_upper = line.strip().upper()
        
        # Check if this line starts a new section
        section_found = False
        for marker, key in sections.items():
            if marker in line_upper:
                current_section = key
                section_found = True
                break
        
        if not section_found and current_section:
            # Add content to current section
            if line.strip() and not line.strip().startswith('**'):
                if result[current_section]:
                    result[current_section] += '\n' + line
                else:
                    result[current_section] = line
    
    # Clean up the text
    for key in result:
        result[key] = result[key].strip()
    
    return result

def create_json_output(scn_data, output_path):
    """Create JSON file with all SCN data."""
    json_data = {
        'metadata': {
            'total_files': len(scn_data),
            'generated_date': time.strftime('%Y-%m-%d %H:%M:%S'),
            'source': 'cobol_src directory'
        },
        'scn_files': scn_data
    }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False)
    
    print(f"\nJSON file created: {output_path}")

def create_excel_output(scn_data, output_path):
    """Create Excel file with SCN data."""
    # Load existing workbook or create new
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
    
    # Remove old tab if exists
    if 'SCN Files Analysis' in wb.sheetnames:
        del wb['SCN Files Analysis']
    
    ws = wb.create_sheet('SCN Files Analysis', 0)
    
    # Headers
    headers = ['File Name', 'File Path', 'Description', 'Menu/Screen Information', 
               'Purpose', 'Calling Programs', 'Number of Callers']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 30  # File Name
    ws.column_dimensions['B'].width = 40  # File Path
    ws.column_dimensions['C'].width = 60  # Description
    ws.column_dimensions['D'].width = 80  # Menu/Screen Information
    ws.column_dimensions['E'].width = 50  # Purpose
    ws.column_dimensions['F'].width = 50  # Calling Programs
    ws.column_dimensions['G'].width = 15  # Number of Callers
    
    # Write data
    for row_idx, item in enumerate(scn_data, start=2):
        ws.cell(row=row_idx, column=1, value=item['filename'])
        ws.cell(row=row_idx, column=2, value=item['path'])
        ws.cell(row=row_idx, column=3, value=item['description'])
        ws.cell(row=row_idx, column=4, value=item['menu_screen_info'])
        ws.cell(row=row_idx, column=5, value=item['purpose'])
        ws.cell(row=row_idx, column=6, value=', '.join(item['calling_programs']))
        ws.cell(row=row_idx, column=7, value=len(item['calling_programs']))
        
        # Wrap text
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    
    wb.save(output_path)
    print(f"Excel file updated: {output_path}")

def main():
    print("=" * 70)
    print("_SCN FILES COMPREHENSIVE ANALYSIS")
    print("=" * 70)
    
    config = load_config()
    if not config:
        return
    
    required_keys = ['AZURE_OPENAI_API_KEY', 'AZURE_OPENAI_ENDPOINT', 'AZURE_OPENAI_DEPLOYMENT']
    missing_keys = [key for key in required_keys if key not in config]
    if missing_keys:
        print(f"Error: Missing required configuration keys: {', '.join(missing_keys)}")
        return
    
    try:
        client = AzureOpenAI(
            api_key=config['AZURE_OPENAI_API_KEY'],
            api_version="2024-02-15-preview",
            azure_endpoint=config['AZURE_OPENAI_ENDPOINT']
        )
        print(f"Connected to Azure OpenAI: {config['AZURE_OPENAI_ENDPOINT']}")
    except Exception as e:
        print(f"Error initializing Azure OpenAI client: {str(e)}")
        return
    
    base_path = os.path.join(os.getcwd(), 'cobol_src')
    excel_path = os.path.join(os.getcwd(), 'scn_files_comprehensive_analysis.xlsx')
    json_path = os.path.join(os.getcwd(), 'scn_files_analysis.json')
    
    # Find all SCN files
    print("\nSearching for _SCN files...")
    scn_files = find_scn_files(base_path)
    print(f"Found {len(scn_files)} _SCN files")
    
    if not scn_files:
        print("No _SCN files found!")
        return
    
    # Process each file
    scn_data = []
    
    for idx, file_info in enumerate(scn_files, start=1):
        print(f"\nProcessing {idx}/{len(scn_files)}: {file_info['filename']}")
        
        # Read file content
        content = read_file_content(file_info['full_path'])
        
        # Get OpenAI analysis
        print(f"  Getting AI analysis...")
        analysis = extract_scn_analysis(
            client, 
            config['AZURE_OPENAI_DEPLOYMENT'],
            file_info['path'], 
            content
        )
        
        # Parse analysis
        parsed = parse_analysis(analysis)
        
        # Find calling programs
        print(f"  Finding calling programs...")
        calling_programs = find_calling_programs(file_info['filename'], base_path)
        
        scn_data.append({
            'filename': file_info['filename'],
            'path': file_info['path'],
            'description': parsed['description'],
            'menu_screen_info': parsed['menu_screen_info'],
            'purpose': parsed['purpose'],
            'calling_programs': calling_programs,
            'caller_count': len(calling_programs)
        })
        
        # Preview
        desc_preview = parsed['description'][:80] + "..." if len(parsed['description']) > 80 else parsed['description']
        print(f"  Description: {desc_preview}")
        print(f"  Calling programs: {len(calling_programs)}")
        
        time.sleep(0.5)
    
    # Create outputs
    print("\n" + "=" * 70)
    print("Creating output files...")
    create_json_output(scn_data, json_path)
    create_excel_output(scn_data, excel_path)
    
    # Summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"Total _SCN files processed: {len(scn_data)}")
    print(f"JSON file: {json_path}")
    print(f"Excel file: {excel_path}")
    print(f"Excel tab: 'SCN Files Analysis'")
    print("=" * 70)

if __name__ == "__main__":
    main()
