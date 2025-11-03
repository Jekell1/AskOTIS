"""
Analyze all menu-related files in cobol_src and create Excel report.
"""
import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def find_menu_files(root_dir):
    """Find all files with 'menu' in the name (case-insensitive)."""
    menu_files = []
    for root, dirs, files in os.walk(root_dir):
        for file in files:
            if 'menu' in file.lower():
                full_path = os.path.join(root, file)
                menu_files.append(full_path)
    return sorted(menu_files)

def extract_description(file_path):
    """Extract description from COBOL file comments."""
    try:
        with open(file_path, 'r', encoding='latin-1') as f:
            content = f.read(5000)  # Read first 5000 chars
            
        # Look for DESC: or DESCRIPTION: in comments
        desc_match = re.search(r'\*\s*DESC(?:RIPTION)?:\s*(.+?)(?:\*-+|\n\s*\*\s*REV:|\n\s*\*{20,})', 
                              content, re.IGNORECASE | re.DOTALL)
        if desc_match:
            desc = desc_match.group(1).strip()
            # Clean up the description
            desc = re.sub(r'\s*\n\s*\*\s*', ' ', desc)
            desc = re.sub(r'\s+', ' ', desc)
            return desc[:500]  # Limit length
        
        # Look for comment block after PROGRAM-ID
        prog_match = re.search(r'PROGRAM-ID\.\s+\w+.*?\n((?:\s*\*[^\n]*\n){2,})', 
                              content, re.IGNORECASE)
        if prog_match:
            comments = prog_match.group(1)
            # Extract meaningful text
            lines = [line.strip('* \t\n') for line in comments.split('\n')]
            lines = [l for l in lines if l and not l.startswith('-') and not l.startswith('*')]
            if lines:
                return ' '.join(lines[:3])[:500]
                
        return "No description found"
    except Exception as e:
        return f"Error reading file: {str(e)}"

def find_menu_copy_files(file_path):
    """Find COPY statements for menu-related .CPY files."""
    try:
        with open(file_path, 'r', encoding='latin-1') as f:
            content = f.read()
        
        # Find COPY statements with 'menu' in the filename
        copy_pattern = r'COPY\s+"([^"]*[Mm][Ee][Nn][Uu][^"]*\.CPY)"'
        matches = re.findall(copy_pattern, content, re.IGNORECASE)
        
        return ', '.join(matches) if matches else "None"
    except Exception as e:
        return f"Error: {str(e)}"

def find_called_programs(file_path):
    """Find programs called by CALL statements."""
    try:
        with open(file_path, 'r', encoding='latin-1') as f:
            content = f.read()
        
        # Find CALL statements
        # Pattern 1: CALL "PROGRAM"
        calls1 = re.findall(r'CALL\s+"([A-Z0-9]+)"', content)
        # Pattern 2: CALL FORM-PROGX (look for what gets moved to FORM-NAM or FORM-PROGX)
        calls2 = re.findall(r'MOVE\s+"([A-Z0-9/]+)"\s+TO\s+FORM-NAM', content)
        
        all_calls = calls1 + calls2
        # Remove duplicates and sort
        unique_calls = sorted(set(all_calls))
        
        return ', '.join(unique_calls) if unique_calls else "None"
    except Exception as e:
        return f"Error: {str(e)}"

def extract_screen_options(file_path):
    """Extract menu options from _SCN.CPY files."""
    try:
        with open(file_path, 'r', encoding='latin-1') as f:
            content = f.read()
        
        options = []
        
        # Pattern for LABEL LINE statements with menu text
        # Look for numbered options or menu items
        label_pattern = r'LABEL\s+LINE\s+\d+\s+COL\s+\d+\s*\n\s*"([^"]+)"'
        matches = re.findall(label_pattern, content, re.IGNORECASE)
        
        for match in matches:
            text = match.strip()
            # Filter for likely menu options (contain numbers, selection prompts, or menu keywords)
            if (re.search(r'^\d+\.', text) or  # Numbered options
                'MENU' in text.upper() or
                'SELECT' in text.upper() or
                re.search(r'^\d{1,2}\s+[A-Z]', text) or  # Number followed by text
                'F\d+' in text):  # Function keys
                options.append(text)
        
        # Also look for the title
        title_match = re.search(r'LABEL\s+LINE\s+0?\d\s+COL\s+\d+\s*\n\s*"([^"]*MENU[^"]*)"', 
                               content, re.IGNORECASE)
        if title_match:
            title = f"[TITLE: {title_match.group(1).strip()}]"
            options.insert(0, title)
        
        return '\n'.join(options) if options else "No menu options found"
    except Exception as e:
        return f"Error: {str(e)}"

def create_excel_report(menu_files, output_file):
    """Create Excel spreadsheet with menu file analysis."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu Files"
    
    # Define headers
    headers = [
        "File Name",
        "File Type",
        "Full Path",
        "Description",
        "Menu Copy Files (CBL only)",
        "Programs Called",
        "Screen Options (SCN only)"
    ]
    
    # Write headers with formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Process each file
    row_num = 2
    for file_path in menu_files:
        path_obj = Path(file_path)
        file_name = path_obj.name
        file_ext = path_obj.suffix.upper()
        
        print(f"Processing: {file_name}")
        
        # Column A: File Name
        ws.cell(row=row_num, column=1, value=file_name)
        
        # Column B: File Type
        ws.cell(row=row_num, column=2, value=file_ext.lstrip('.'))
        
        # Column C: Full Path
        ws.cell(row=row_num, column=3, value=file_path)
        
        # Column D: Description (for .CBL and .CPY files)
        if file_ext in ['.CBL', '.CPY', '.COB']:
            description = extract_description(file_path)
            ws.cell(row=row_num, column=4, value=description)
        else:
            ws.cell(row=row_num, column=4, value="N/A")
        
        # Column E: Menu Copy Files (only for .CBL files)
        if file_ext == '.CBL':
            menu_copies = find_menu_copy_files(file_path)
            ws.cell(row=row_num, column=5, value=menu_copies)
        else:
            ws.cell(row=row_num, column=5, value="N/A")
        
        # Column F: Programs Called (for .CBL files)
        if file_ext == '.CBL':
            called_programs = find_called_programs(file_path)
            ws.cell(row=row_num, column=6, value=called_programs)
        else:
            ws.cell(row=row_num, column=6, value="N/A")
        
        # Column G: Screen Options (only for _SCN files)
        if '_SCN' in file_name.upper():
            screen_options = extract_screen_options(file_path)
            cell = ws.cell(row=row_num, column=7, value=screen_options)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            ws.cell(row=row_num, column=7, value="N/A")
        
        # Format row
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        row_num += 1
    
    # Adjust column widths
    column_widths = {
        'A': 30,  # File Name
        'B': 10,  # File Type
        'C': 60,  # Full Path
        'D': 50,  # Description
        'E': 30,  # Menu Copy Files
        'F': 40,  # Programs Called
        'G': 60   # Screen Options
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    wb.save(output_file)
    print(f"\nExcel report saved to: {output_file}")
    print(f"Total menu files analyzed: {len(menu_files)}")

def main():
    """Main function."""
    cobol_src = Path("cobol_src")
    
    if not cobol_src.exists():
        print(f"Error: Directory '{cobol_src}' not found!")
        return
    
    print("Scanning for menu files...")
    menu_files = find_menu_files(cobol_src)
    print(f"Found {len(menu_files)} menu-related files\n")
    
    output_file = "menu_files_analysis.xlsx"
    create_excel_report(menu_files, output_file)

if __name__ == "__main__":
    main()
